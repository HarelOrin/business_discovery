"""
NAICS Source Puller — Ingest official Census Bureau 2022 NAICS files
into raw_intake_naics.json.

Reads:
  data/raw/naics/2-6_digit_2022_Codes.xlsx   (codes + titles)
  data/raw/naics/2022_NAICS_Descriptions.xlsx (codes + titles + descriptions)

Emits:
  data/output/raw_intake_naics.json
"""

import json
import os
from datetime import datetime, timezone

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
RAW_DIR = os.path.join(BASE_DIR, "data", "raw", "naics")
OUT_DIR = os.path.join(BASE_DIR, "data", "output")


def load_codes(path: str) -> dict[str, str]:
    """Load NAICS codes and titles from the codes xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    codes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, code, title = row[0], row[1], row[2]
        if code is None or title is None:
            continue
        code_str = str(code).strip()
        if not code_str or not code_str.isdigit():
            continue
        codes[code_str] = title.strip()

    wb.close()
    return codes


def load_descriptions(path: str) -> dict[str, str]:
    """Load NAICS descriptions from the descriptions xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    descriptions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, title, desc = row[0], row[1], row[2]
        if code is None:
            continue
        code_str = str(code).strip()
        if not code_str:
            continue
        desc_text = str(desc).strip() if desc else ""
        descriptions[code_str] = desc_text

    wb.close()
    return descriptions


LEVEL_NAMES = {
    2: "sector",
    3: "subsector",
    4: "industry_group",
    5: "naics_industry",
    6: "national_industry",
}


def derive_parent_code(code: str) -> str | None:
    """Derive parent NAICS code by trimming the last digit."""
    if len(code) <= 2:
        return None
    return code[:-1]


def build_records(codes: dict[str, str], descriptions: dict[str, str]) -> list[dict]:
    """Build raw intake records from codes and descriptions."""
    ingested_at = datetime.now(timezone.utc).isoformat()
    records = []

    for code, title in sorted(codes.items(), key=lambda x: x[0]):
        code_len = len(code)
        if code_len < 2 or code_len > 6:
            continue

        # Clean title: some description-file titles have trailing 'T'
        clean_title = title.rstrip("T").strip() if title.endswith("T") else title

        desc = descriptions.get(code, "")

        record = {
            "naics_code": code,
            "title": clean_title,
            "description": desc if desc else None,
            "hierarchy_level": code_len,
            "hierarchy_label": LEVEL_NAMES.get(code_len, "unknown"),
            "parent_code": derive_parent_code(code),
            "source_family": "naics",
            "source_ref": f"census.gov/naics/2022NAICS code={code}",
            "ingested_at": ingested_at,
        }
        records.append(record)

    return records


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    codes_path = os.path.join(RAW_DIR, "2-6_digit_2022_Codes.xlsx")
    desc_path = os.path.join(RAW_DIR, "2022_NAICS_Descriptions.xlsx")

    print(f"Loading NAICS codes from {codes_path}")
    codes = load_codes(codes_path)
    print(f"  -> {len(codes)} codes loaded")

    print(f"Loading NAICS descriptions from {desc_path}")
    descriptions = load_descriptions(desc_path)
    print(f"  -> {len(descriptions)} descriptions loaded")

    print("Building records...")
    records = build_records(codes, descriptions)
    print(f"  -> {len(records)} records built")

    # Breakdown by hierarchy level
    from collections import Counter
    level_counts = Counter(r["hierarchy_level"] for r in records)
    for level in sorted(level_counts):
        label = LEVEL_NAMES.get(level, "unknown")
        print(f"     Level {level} ({label}): {level_counts[level]}")

    # Check description coverage
    with_desc = sum(1 for r in records if r["description"])
    print(f"  -> {with_desc}/{len(records)} have descriptions ({100*with_desc/len(records):.1f}%)")

    # Check for null titles
    null_titles = sum(1 for r in records if not r["title"])
    print(f"  -> {null_titles} records with null/empty title")

    out_path = os.path.join(OUT_DIR, "raw_intake_naics.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
    print(f"\nWritten to {out_path}")
    print(f"File size: {os.path.getsize(out_path):,} bytes")

    # Print sample records
    print("\nSample records:")
    for r in records[:3]:
        print(json.dumps(r, indent=2, ensure_ascii=False)[:500])
        print("---")

    return records


if __name__ == "__main__":
    main()
